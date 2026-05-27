"""Convert an .xlsx workbook into one TSV per sheet, preserving raw layout.

These workbooks (historical Ainu old-documents, dialect database) contain
many heterogeneous sheets with bespoke column layouts. Rather than try to
normalize them, dump each sheet verbatim as TSV and let downstream consumers
read them with their original schema. A metadata.yaml records the source
file, conversion date, and sheet-by-sheet row/col counts.
"""

from __future__ import annotations

import argparse
import csv
import datetime
import json
import sys
from pathlib import Path

import openpyxl  # type: ignore[import-not-found]


def safe_sheet_name(name: str) -> str:
    """Filesystem-safe sheet filename."""
    keep = [c if c.isalnum() or c in "-_." else "_" for c in name]
    return "".join(keep) or "sheet"


def convert(xlsx_path: Path, out_dir: Path) -> dict:
    out_dir.mkdir(parents=True, exist_ok=True)
    sheets_dir = out_dir / "sheets"
    sheets_dir.mkdir(exist_ok=True)
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    sheet_info = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        # Trim entirely empty trailing rows.
        while rows and all(c is None or c == "" for c in rows[-1]):
            rows.pop()
        if not rows:
            sheet_info.append({"sheet": sheet_name, "rows": 0, "cols": 0, "file": None})
            continue
        max_cols = max(len(r) for r in rows)
        file_stem = safe_sheet_name(sheet_name)
        # Avoid collisions when two sheets sanitize to the same stem.
        out_path = sheets_dir / f"{file_stem}.tsv"
        suffix = 1
        while out_path.exists():
            out_path = sheets_dir / f"{file_stem}_{suffix}.tsv"
            suffix += 1
        with out_path.open("w", encoding="utf-8", newline="") as fh:
            writer = csv.writer(fh, delimiter="\t", lineterminator="\n")
            for row in rows:
                writer.writerow([
                    "" if v is None else str(v).replace("\t", " ").replace("\r", " ")
                    for v in row + (None,) * (max_cols - len(row))
                ])
        sheet_info.append(
            {
                "sheet": sheet_name,
                "rows": len(rows),
                "cols": max_cols,
                "file": str(out_path.relative_to(out_dir)),
            }
        )
    wb.close()

    manifest = {
        "source_xlsx": str(xlsx_path),
        "converted_at": datetime.datetime.now().strftime("%Y-%m-%d"),
        "sheet_count": len(sheet_info),
        "sheets": sheet_info,
    }
    (out_dir / "manifest.json").write_text(
        json.dumps(manifest, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    return manifest


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("xlsx", type=Path)
    parser.add_argument("--out", type=Path, required=True)
    args = parser.parse_args()
    manifest = convert(args.xlsx, args.out)
    print(
        f"converted {args.xlsx} -> {args.out} ({manifest['sheet_count']} sheets, "
        f"{sum(s['rows'] for s in manifest['sheets'])} rows total)"
    )


if __name__ == "__main__":
    sys.exit(main())
