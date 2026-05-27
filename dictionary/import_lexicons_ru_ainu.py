"""Import Lexicons.ru Ainu workbook (Ainu-English-Ainu + comparison sheets)."""

from __future__ import annotations

import csv
import zipfile
import urllib.request
from pathlib import Path

import openpyxl


DICT_ROOT = Path("/home/mkpoli/projects/Ainu/ainu-dictionaries")
FOLDER = DICT_ROOT / "2024_LexiconsRu_Ainu-English-Ainu"
ZIP_URL = "https://lexicons.ru/extinct/a/ainu/_zip/ainu-eng-ainu.zip"


def fetch_zip(path: Path) -> None:
    with urllib.request.urlopen(ZIP_URL, timeout=30) as response, path.open("wb") as out:
        out.write(response.read())


def norm(value: object) -> str:
    return "" if value is None else str(value).strip()


def main() -> None:
    FOLDER.mkdir(parents=True, exist_ok=True)
    zip_path = FOLDER / "ainu-eng-ainu.zip"
    fetch_zip(zip_path)
    with zipfile.ZipFile(zip_path) as zf:
        zf.extractall(FOLDER)
    xlsx_path = FOLDER / "ainu-eng-ainu.xlsx"
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)

    total_rows = 0
    manifest: list[tuple[str, int, int, str]] = []
    sheets_dir = FOLDER / "sheets"
    sheets_dir.mkdir(exist_ok=True)
    for ws in wb.worksheets:
        if ws.title == "Ainu-Eng":
            columns = ["lemma", "meaning_en"]
            start_row = 1
        elif ws.title == "Eng-Ainu":
            columns = ["meaning_en", "lemma"]
            start_row = 1
        elif ws.title == "Ainu-IE":
            columns = ["ainu", "indo_european", "note"]
            start_row = 2
        elif ws.title == "Ainu-Basque":
            columns = ["ainu", "ainu_meaning_en", "basque", "basque_meaning_en", "note"]
            start_row = 3
        else:
            columns = [f"col{i}" for i in range(1, ws.max_column + 1)]
            start_row = 1

        rows = []
        for values in ws.iter_rows(min_row=start_row, values_only=True):
            row = {col: norm(values[idx]) if idx < len(values) else "" for idx, col in enumerate(columns)}
            if any(row.values()):
                rows.append(row)
        out_path = sheets_dir / f"{ws.title}.tsv"
        with out_path.open("w", encoding="utf-8", newline="") as fh:
            writer = csv.DictWriter(fh, fieldnames=columns, delimiter="\t", lineterminator="\n")
            writer.writeheader()
            writer.writerows(rows)
        total_rows += len(rows)
        manifest.append((ws.title, len(rows), ws.max_column, ", ".join(columns)))

    # Main table mirrors Ainu-Eng for dictionary consumers.
    main_rows = list(csv.DictReader((sheets_dir / "Ainu-Eng.tsv").open(encoding="utf-8"), delimiter="\t"))
    with (FOLDER / "original.tsv").open("w", encoding="utf-8", newline="") as fh:
        writer = csv.DictWriter(fh, fieldnames=["lemma", "meaning_en"], delimiter="\t", lineterminator="\n")
        writer.writeheader()
        writer.writerows(main_rows)

    manifest_text = "\n".join(
        f"  - sheet: {name}\n    rows: {rows}\n    columns: {cols}" for name, rows, _ncols, cols in manifest
    )
    (FOLDER / "metadata.yaml").write_text(
        f"""type: workbook
title: Лексика языка айну — Ainu-English-Ainu workbook
title_en: Lexicons.ru Ainu-English-Ainu workbook
author:
  en: Garshin, Igor
  ru: Гаршин, Игорь
year: 2024
url: https://lexicons.ru/extinct/a/ainu/index.html
source: |
  Downloaded from Lexicons.ru. The workbook contains Ainu-English and
  English-Ainu tables matching the RaccoonBend/Shibatani 605-entry list, plus
  Russian-language Indo-European comparison notes and an Ainu-Basque comparison
  sheet. `original.tsv` exposes the Ainu-English sheet; all workbook sheets are
  preserved under `sheets/`.
sheets:
{manifest_text}
columns:
  lemma: Ainu form (from Ainu-Eng sheet)
  meaning_en: English gloss (from Ainu-Eng sheet)
""",
        encoding="utf-8",
    )
    print(f"wrote Lexicons.ru workbook: {total_rows} rows across {len(manifest)} sheets")


if __name__ == "__main__":
    main()
